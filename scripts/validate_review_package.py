from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
REQUIRED = [
    "OPEN_THIS_FIRST.html",
    "README.md",
    "recruiter_summary.md",
    "AI_REVIEW_README.md",
    "DATA_ACCESS.md",
    "validation/validation_report.md",
    "validation/validation_checks.csv",
    "validation/artifact_manifest_integrity.csv",
    "excel/Project5_FraudOperationalRisk_Model.xlsx",
    "reports/fraud_risk_committee_memo.md",
    "reports/operational_risk_committee_memo.md",
    "data_contract/source_manifest.csv",
    "data/observed/fraud_pca_sample.csv",
    "data/synthetic/synthetic_transactions_sample.csv",
    "data/review_samples/hybrid_decisions_sample.csv",
    "data/review_samples/alert_queue_sample.csv",
]


def main() -> None:
    failures = []
    for relative in REQUIRED:
        path = ROOT / relative
        if not path.exists() or path.stat().st_size == 0:
            failures.append(f"Missing or empty: {relative}")

    html_path = ROOT / "OPEN_THIS_FIRST.html"
    if html_path.exists():
        content = html_path.read_text(encoding="utf-8")
        links = re.findall(r'href="([^"]+)"', content)
        local = [link for link in links if not link.startswith(("http://", "https://", "#"))]
        for link in local:
            if not (ROOT / link).exists():
                failures.append(f"Unresolved HTML link: {link}")

    workbook_path = ROOT / "excel/Project5_FraudOperationalRisk_Model.xlsx"
    if workbook_path.exists():
        workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
        if len(workbook.sheetnames) != 14:
            failures.append(f"Excel sheet count is {len(workbook.sheetnames)}, expected 14")

    validation_path = ROOT / "validation/validation_checks.csv"
    if validation_path.exists():
        validation = pd.read_csv(validation_path)
        if not validation["status"].eq("PASS").all():
            failures.append("Included final validation contains unresolved FAIL")

    source_manifest = ROOT / "data_contract/source_manifest.csv"
    if source_manifest.exists():
        source = pd.read_csv(source_manifest)
        required_columns = {"dataset_id", "source_url", "licence_or_access_terms", "canonical_path", "packaged_path", "downstream_use"}
        if not required_columns.issubset(source.columns):
            failures.append("Source manifest lacks public lineage columns")

    release_path = ROOT / "testing/release_readiness_checklist.csv"
    if release_path.exists():
        release = pd.read_csv(release_path)
        if "actual_organisational_approval" not in release.columns or not release["actual_organisational_approval"].eq("NO").all():
            failures.append("Release checklist does not clearly mark organisational approval as NO")

    drive_pattern = re.compile(r"(?<![A-Za-z])[A-Za-z]:[\\/]")
    public_ext = {".py", ".md", ".csv", ".json", ".html", ".txt", ".log"}
    path_hits = []
    for path in ROOT.rglob("*"):
        if path.is_file() and path.suffix.lower() in public_ext and "__pycache__" not in path.parts:
            content = path.read_text(encoding="utf-8", errors="ignore")
            if drive_pattern.search(content):
                path_hits.append(path.relative_to(ROOT).as_posix())
    if path_hits:
        failures.append(f"Absolute personal paths found: {sorted(set(path_hits))}")

    if failures:
        print("AI REVIEW PACKAGE VALIDATION FAIL")
        print("\n".join(failures))
        raise SystemExit(1)
    print("AI REVIEW PACKAGE VALIDATION PASS")


if __name__ == "__main__":
    main()
