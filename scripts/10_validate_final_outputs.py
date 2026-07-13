from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

from _project5_common import ROOT, add_metadata, dataframe_to_markdown, file_sha256, get_logger, write_csv, write_markdown


REQUIRED_FILES = [
    "README.md",
    "OPEN_THIS_FIRST.html",
    "business_question.md",
    "recruiter_summary.md",
    "limitation_note.md",
    "change_log.md",
    "requirements.txt",
    "DATA_ACCESS.md",
    "AI_REVIEW_README.md",
    "data_contract/p5_observed_input_contract.csv",
    "data_contract/p5_synthetic_input_contract.csv",
    "data_contract/p5_data_availability_check.md",
    "data_contract/observed_vs_synthetic_design.md",
    "data_contract/population_reconciliation.csv",
    "data_contract/observed_split_reconciliation.csv",
    "data_contract/synthetic_population_reconciliation.csv",
    "data_contract/source_manifest.csv",
    "data/observed/fraud_pca_input.csv.gz",
    "data/observed/fraud_pca_sample.csv",
    "data/synthetic/synthetic_transactions.csv.gz",
    "data/synthetic/synthetic_transactions_sample.csv",
    "data/synthetic/synthetic_customers.csv",
    "data/synthetic/synthetic_devices.csv",
    "data/synthetic/synthetic_beneficiaries.csv",
    "data/synthetic/synthetic_disputes.csv",
    "data/reference/p3_customer_risk_context.csv",
    "data/review_samples/synthetic_customers_sample.csv",
    "data/review_samples/synthetic_devices_sample.csv",
    "data/review_samples/synthetic_beneficiaries_sample.csv",
    "data/review_samples/synthetic_disputes_sample.csv",
    "data/review_samples/hybrid_decisions_sample.csv",
    "data/review_samples/alert_queue_sample.csv",
    "methodology/fraud_label_definition.md",
    "methodology/synthetic_fraud_generation_methodology.md",
    "methodology/model_methodology.md",
    "methodology/anomaly_detection_methodology.md",
    "methodology/threshold_economics_methodology.md",
    "methodology/threshold_economics_decision.md",
    "methodology/probability_calibration_decision.md",
    "methodology/friendly_fraud_claim_methodology.md",
    "methodology/fraud_rule_methodology.md",
    "methodology/hybrid_decision_engine_methodology.md",
    "methodology/alert_prioritisation_methodology.md",
    "methodology/operational_segment_fairness_spec.md",
    "models/model_comparison.csv",
    "models/champion_model_card.md",
    "models/challenger_model_card.md",
    "models/threshold_analysis.csv",
    "models/calibration_summary.csv",
    "models/feature_governance.csv",
    "models/selected_observed_threshold.json",
    "models/model_uncertainty_intervals.csv",
    "models/synthetic_model_summary.csv",
    "models/synthetic_hybrid_thresholds.csv",
    "models/synthetic_p3_context_comparison.csv",
    "rules/fraud_rule_inventory.csv",
    "rules/rule_priority_matrix.csv",
    "rules/decision_matrix.csv",
    "rules/reason_code_dictionary.csv",
    "rules/override_policy.csv",
    "outputs/observed_fraud_kpi_summary.csv",
    "outputs/observed_model_predictions.csv.gz",
    "outputs/synthetic_transaction_scores.csv.gz",
    "outputs/hybrid_decision_output.csv.gz",
    "outputs/alert_queue.csv.gz",
    "outputs/alert_outcome_summary.csv",
    "outputs/fraud_amount_capture_curve.csv",
    "outputs/threshold_economics.csv",
    "outputs/rule_performance_summary.csv",
    "outputs/segment_risk_summary.csv",
    "outputs/control_effectiveness_summary.csv",
    "outputs/hybrid_incremental_value.csv",
    "outputs/friendly_fraud_claim_risk_summary.csv",
    "outputs/priority_queue_efficiency.csv",
    "operational_risk/incident_register.csv",
    "operational_risk/loss_event_summary.csv",
    "operational_risk/root_cause_analysis.md",
    "operational_risk/scenario_analysis.csv",
    "operational_risk/control_failure_library.csv",
    "governance/assumption_register.csv",
    "governance/fraud_risk_appetite_register.csv",
    "governance/operational_risk_appetite_register.csv",
    "governance/owner_approver_matrix.csv",
    "governance/monitoring_plan.csv",
    "governance/management_action_library.csv",
    "governance/exception_register.csv",
    "governance/model_use_register.csv",
    "governance/audit_evidence_checklist.csv",
    "testing/uat_test_cases.csv",
    "testing/sit_test_cases.csv",
    "testing/negative_test_cases.csv",
    "testing/test_evidence_log.csv",
    "testing/release_readiness_checklist.csv",
    "testing/rollback_plan.md",
    "testing/incident_management_playbook.md",
    "validation/independent_review_checklist.csv",
    "validation/model_validation_summary.md",
    "validation/control_validation_summary.md",
    "validation/operational_segment_fairness.csv",
    "validation/operational_segment_fairness_actions.csv",
    "reports/observed_fraud_eda.md",
    "reports/model_performance_report.md",
    "reports/probability_calibration_report.md",
    "reports/threshold_recommendation_memo.md",
    "reports/control_effectiveness_report.md",
    "reports/alert_operations_report.md",
    "reports/incident_management_report.md",
    "reports/fraud_risk_committee_memo.md",
    "reports/operational_risk_committee_memo.md",
    "reports/operational_segment_fairness_note.md",
    "reports/project3_project6_linkage_note.md",
    "excel/Project5_FraudOperationalRisk_Model.xlsx",
    "scripts/01_load_validate_observed_data.py",
    "scripts/02_build_observed_fraud_benchmark.py",
    "scripts/03_generate_synthetic_control_data.py",
    "scripts/04_build_synthetic_features_and_labels.py",
    "scripts/05_build_fraud_rules.py",
    "scripts/06_build_hybrid_decision_engine.py",
    "scripts/07_build_alert_operations.py",
    "scripts/08_build_operational_risk_outputs.py",
    "scripts/09_generate_excel_and_reports.py",
    "scripts/10_validate_final_outputs.py",
    "scripts/run_full_pipeline.py",
    "scripts/package_project.py",
    "scripts/validate_review_package.py",
]


PUBLIC_SCAN_EXTENSIONS = {
    ".py", ".md", ".csv", ".json", ".html", ".txt", ".log"
}


def build_artifact_manifest() -> pd.DataFrame:
    excluded = {
        "validation/artifact_manifest.csv",
        "validation/artifact_manifest_integrity.csv",
    }
    generated_files = [
        path
        for path in ROOT.rglob("*")
        if path.is_file()
        and "__pycache__" not in path.parts
        and path.suffix.lower() != ".pyc"
        and path.relative_to(ROOT).as_posix() not in excluded
    ]
    manifest_rows = []
    for path in sorted(generated_files):
        manifest_rows.append(
            {
                "relative_path": path.relative_to(ROOT).as_posix(),
                "size_bytes": path.stat().st_size,
                "sha256": file_sha256(path),
                "artifact_type": path.suffix.lower() or "none",
            }
        )
    return add_metadata(pd.DataFrame(manifest_rows), "Mixed", "Derived", "Project filesystem", "Artifact integrity manifest")


def write_manifest_integrity(manifest: pd.DataFrame) -> pd.DataFrame:
    rows = []
    seen: set[str] = set()
    for _, row in manifest.iterrows():
        relative = str(row["relative_path"])
        path = ROOT / relative
        actual_sha = file_sha256(path) if path.exists() else ""
        rows.append(
            {
                "relative_path": relative,
                "exists": bool(path.exists()),
                "duplicate_path": relative in seen,
                "recorded_sha256": row["sha256"],
                "actual_sha256": actual_sha,
                "hash_match": bool(path.exists()) and actual_sha == row["sha256"],
                "manifest_self_reference": relative == "validation/artifact_manifest.csv",
                "status": "PASS"
                if path.exists() and actual_sha == row["sha256"] and relative not in seen and relative != "validation/artifact_manifest.csv"
                else "FAIL",
            }
        )
        seen.add(relative)
    integrity = add_metadata(
        pd.DataFrame(rows),
        "Mixed",
        "Derived",
        "Post-manifest hash verification",
        "Artifact integrity evidence; manifest excludes itself",
    )
    write_csv(integrity, "validation/artifact_manifest_integrity.csv")
    return integrity


def absolute_path_matches() -> list[str]:
    drive_pattern = re.compile(r"(?<![A-Za-z])[A-Za-z]:[\\/]")
    matches = []
    for path in ROOT.rglob("*"):
        if not path.is_file() or "__pycache__" in path.parts:
            continue
        if path.suffix.lower() not in PUBLIC_SCAN_EXTENSIONS:
            continue
        try:
            content = path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        if drive_pattern.search(content):
            matches.append(path.relative_to(ROOT).as_posix())
    return sorted(set(matches))


def main() -> None:
    start_ts = datetime.now().isoformat(timespec="seconds")
    checks: list[dict] = []

    def add(group: str, check_id: str, description: str, passed: bool, evidence: str, criticality: str = "Critical", owner: str = "Project Owner", action: str = "None") -> None:
        checks.append(
            {
                "check_id": check_id,
                "check_group": group,
                "criticality": criticality,
                "check_description": description,
                "status": "PASS" if bool(passed) else "FAIL",
                "evidence": evidence,
                "owner": owner,
                "action_if_failed": action if action != "None" else "Correct before release",
            }
        )

    for index, relative in enumerate(REQUIRED_FILES, start=1):
        path = ROOT / relative
        add("Files", f"FILE-{index:03d}", f"Required artifact exists: {relative}", path.exists() and path.stat().st_size > 0, f"exists={path.exists()} size={path.stat().st_size if path.exists() else 0}")

    observed = pd.read_csv(ROOT / "data/observed/fraud_pca_input.csv.gz")
    add("Observed data", "OBS-001", "Observed row count is 284,807", len(observed) == 284_807, f"rows={len(observed)}")
    add("Observed data", "OBS-002", "Observed fraud count is 492", int(observed["fraud_flag"].sum()) == 492, f"fraud={int(observed['fraud_flag'].sum())}")
    add("Observed data", "OBS-003", "Fraud target contains only 0/1", observed["fraud_flag"].isin([0, 1]).all(), str(sorted(observed["fraud_flag"].unique())))
    add("Observed data", "OBS-004", "Amounts are non-negative", observed["amount"].ge(0).all(), f"min={observed['amount'].min()}")
    add("Observed data", "OBS-005", "Transaction IDs are unique", not observed["transaction_id"].duplicated().any(), f"duplicates={observed['transaction_id'].duplicated().sum()}")
    add("Observed data", "OBS-006", "No missing model features", not observed[[f"v{i}" for i in range(1, 29)] + ["amount"]].isna().any().any(), "v1-v28 and amount")
    split = pd.read_csv(ROOT / "data_contract/observed_split_reconciliation.csv")
    add("Observed data", "OBS-007", "Split rows reconcile", int(split["rows"].sum()) == len(observed), f"split_rows={int(split['rows'].sum())}")
    add("Observed data", "OBS-008", "Train/validation/test all contain fraud", (split["fraud_count"] > 0).all(), split[["split", "fraud_count"]].to_json(orient="records"))

    model = pd.read_csv(ROOT / "models/model_comparison.csv")
    champion_test = model.loc[model["population"].eq("test") & model["model_role"].eq("Champion")]
    add("Model metrics", "MOD-001", "Exactly one champion on test", len(champion_test) == 1, f"count={len(champion_test)}")
    add("Model metrics", "MOD-002", "Champion test PR-AUC exceeds prevalence materially", champion_test.iloc[0]["pr_auc"] > champion_test.iloc[0]["fraud_rate"] * 10, f"pr_auc={champion_test.iloc[0]['pr_auc']:.4f}")
    add("Model metrics", "MOD-003", "Logistic transparent baseline present", model["model"].eq("logistic_no_resampling").any(), "model_comparison.csv")
    add("Model metrics", "MOD-004", "Tree challenger present", model["model_family"].eq("tree").any(), "model_comparison.csv")
    add("Model metrics", "MOD-005", "Anomaly benchmark present", model["model_family"].eq("anomaly").any(), "model_comparison.csv")
    add("Model metrics", "MOD-006", "SMOTE train-only comparator present", model["resampling_strategy"].eq("SMOTE_train_only").any(), "model_comparison.csv")
    add("Model metrics", "MOD-007", "All probabilities/metrics finite", np.isfinite(model[["pr_auc", "roc_auc", "precision", "recall", "brier_score"]].to_numpy()).all(), "model comparison numeric fields")

    calibration = pd.read_csv(ROOT / "models/calibration_summary.csv")
    add("Calibration", "CAL-001", "Platt calibrated test result exists", ((calibration["population"] == "test") & (calibration["calibration_stage"] == "platt_calibrated")).any(), "calibration_summary.csv")
    add("Calibration", "CAL-002", "Isotonic not used below 200-event gate", not calibration["calibration_method"].astype(str).str.contains("isotonic", case=False).any(), "validation fraud count below gate")
    add("Calibration", "CAL-003", "Test not used for calibration fitting", calibration.loc[calibration["population"].eq("test"), "acceptance_note"].str.contains("no test fitting", case=False).all(), "acceptance_note")

    economics = pd.read_csv(ROOT / "outputs/threshold_economics.csv")
    recommendation = economics.loc[economics["population"].eq("validation") & economics["recommended_threshold_flag"].eq(1)]
    add("Threshold economics", "THR-001", "One validation recommendation", len(recommendation) == 1, f"count={len(recommendation)}")
    add("Threshold economics", "THR-002", "Recommended validation threshold within capacity", recommendation.iloc[0]["capacity_status"] == "PASS", str(recommendation.iloc[0]["alerts_per_monitoring_period_proxy"]))
    add("Threshold economics", "THR-003", "Recommended validation net benefit positive", recommendation.iloc[0]["realised_net_benefit"] > 0, str(recommendation.iloc[0]["realised_net_benefit"]))
    add("Threshold economics", "THR-004", "Candidate grid frozen and complete", set(np.round(economics["target_alert_rate"].unique(), 4)) == {0.0005, 0.001, 0.002, 0.005, 0.01, 0.02, 0.05}, str(sorted(economics["target_alert_rate"].unique())))
    add("Threshold economics", "THR-005", "Backtest and prospective economics separated", {"realised_prevented_loss_proxy", "prospective_expected_prevented_loss_proxy"}.issubset(economics.columns), "separate columns")

    synthetic = pd.read_csv(ROOT / "data/synthetic/synthetic_transactions.csv.gz")
    add("Synthetic data", "SYN-001", "Synthetic transaction row count is 250,000", len(synthetic) == 250_000, f"rows={len(synthetic)}")
    add("Synthetic data", "SYN-002", "Synthetic IDs unique", not synthetic["transaction_id"].duplicated().any(), f"duplicates={synthetic['transaction_id'].duplicated().sum()}")
    add("Synthetic data", "SYN-003", "Synthetic fraud label is 0/1", synthetic["transaction_fraud_label"].isin([0, 1]).all(), str(sorted(synthetic["transaction_fraud_label"].unique())))
    add("Synthetic data", "SYN-004", "Synthetic fraud rate non-degenerate", 0.002 < synthetic["transaction_fraud_label"].mean() < 0.05, f"rate={synthetic['transaction_fraud_label'].mean():.4%}")
    add("Synthetic data", "SYN-005", "All fraud archetypes represented", synthetic.loc[synthetic["transaction_fraud_label"].eq(1), "fraud_type"].nunique() >= 6, f"types={synthetic.loc[synthetic['transaction_fraud_label'].eq(1),'fraud_type'].nunique()}")
    add("Synthetic data", "SYN-006", "No hidden latent columns exposed in final transactions", not any(column.startswith("hidden_") for column in synthetic.columns), "final synthetic_transactions.csv.gz")
    disputes = pd.read_csv(ROOT / "data/synthetic/synthetic_disputes.csv")
    add("Synthetic data", "SYN-007", "Friendly fraud claim layer implemented", len(disputes) >= 10_000 and "friendly_fraud_label" in disputes, f"disputes={len(disputes)}")
    add("Synthetic data", "SYN-008", "Friendly and transaction labels stored separately", {"friendly_fraud_label", "transaction_fraud_label"}.issubset(disputes.columns), "synthetic_disputes.csv")
    synthetic_model = pd.read_csv(ROOT / "models/synthetic_model_summary.csv")
    synthetic_test = synthetic_model.loc[
        synthetic_model["population"].eq("test") & synthetic_model["score"].eq("platt_calibrated")
    ].iloc[0]
    add("Synthetic data", "SYN-009", "Synthetic model PR-AUC is at least five times prevalence", synthetic_test["pr_auc"] >= synthetic_test["observed_fraud_rate"] * 5, f"pr_auc={synthetic_test['pr_auc']:.4f} prevalence={synthetic_test['observed_fraud_rate']:.4f}")
    add("Synthetic data", "SYN-010", "Synthetic calibrated mean aligns within 20% relative prevalence gap", abs(synthetic_test["mean_score"] - synthetic_test["observed_fraud_rate"]) / synthetic_test["observed_fraud_rate"] <= 0.20, f"mean={synthetic_test['mean_score']:.4f} observed={synthetic_test['observed_fraud_rate']:.4f}")

    anti = pd.read_csv(ROOT / "validation/synthetic_anti_circularity_tests.csv")
    add("Anti-circularity", "ANTI-001", "All anti-circularity tests pass", anti["status"].eq("PASS").all(), anti[["test_id", "status"]].to_json(orient="records"))
    add("Anti-circularity", "ANTI-002", "Rules do not have 100% aggregate precision", float(anti.loc[anti["test_id"].eq("obvious_rule_precision_below_100pct"), "value"].iloc[0]) < 1, "obvious rule precision")
    add("Anti-circularity", "ANTI-003", "Rules do not have 100% aggregate recall", float(anti.loc[anti["test_id"].eq("obvious_rule_recall_below_100pct"), "value"].iloc[0]) < 1, "obvious rule recall")

    rules = pd.read_csv(ROOT / "rules/fraud_rule_inventory.csv")
    add("Rule inventory", "RULE-001", "At least 15 governed rules", len(rules) >= 15, f"rules={len(rules)}")
    add("Rule inventory", "RULE-002", "Rule IDs unique", not rules["rule_id"].duplicated().any(), "fraud_rule_inventory.csv")
    add("Rule inventory", "RULE-003", "Every rule has reason and action", rules[["reason_code", "action"]].notna().all().all(), "reason_code/action")
    add("Rule inventory", "RULE-004", "No PCA variable in rule conditions", not rules["condition"].str.contains(r"\b[vV](?:[1-9]|1[0-9]|2[0-8])\b", regex=True).any(), "rule condition scan")
    reasons = pd.read_csv(ROOT / "rules/reason_code_dictionary.csv")
    add("Rule inventory", "RULE-005", "No PCA names in reason codes", not reasons.astype(str).apply(lambda col: col.str.contains(r"\b[vV](?:[1-9]|1[0-9]|2[0-8])\b", regex=True)).any().any(), "reason dictionary scan")

    decisions = pd.read_csv(ROOT / "outputs/hybrid_decision_output.csv.gz")
    allowed_decisions = {"ALLOW", "STEP_UP", "HOLD", "MANUAL_REVIEW", "DECLINE", "BLOCK_ACCOUNT", "DATA_EXCEPTION"}
    add("Decision engine", "DEC-001", "All decision values allowed", set(decisions["decision"].unique()).issubset(allowed_decisions), str(sorted(decisions["decision"].unique())))
    add("Decision engine", "DEC-002", "All decisions have primary reason", decisions["primary_reason_code"].notna().all(), "primary_reason_code")
    add("Decision engine", "DEC-003", "All scores within 0-1", decisions["synthetic_model_score"].between(0, 1).all(), "synthetic_model_score")
    incremental = pd.read_csv(ROOT / "outputs/hybrid_incremental_value.csv")
    add("Decision engine", "DEC-004", "Required component comparisons present", {"Model only", "Rules only", "Anomaly only", "Model + rules", "Model + rules + anomaly", "Final hybrid actionable"}.issubset(set(incremental["component_set"])), "hybrid_incremental_value.csv")
    add("Decision engine", "DEC-005", "Model or rules show unique contribution", incremental["unique_fraud_contribution"].max() > 0, f"max_unique={incremental['unique_fraud_contribution'].max()}")
    final_hybrid = incremental.loc[incremental["component_set"].eq("Final hybrid actionable")].iloc[0]
    add("Decision engine", "DEC-006", "Final hybrid precision is at least 5%", final_hybrid["precision"] >= 0.05, f"precision={final_hybrid['precision']:.2%}")
    add("Decision engine", "DEC-007", "Final hybrid net benefit proxy is positive", final_hybrid["net_benefit_proxy"] > 0, f"net_benefit={final_hybrid['net_benefit_proxy']:.2f}")
    add("Decision engine", "DEC-008", "Final hybrid alert rate is at most 3%", final_hybrid["alert_rate"] <= 0.03, f"alert_rate={final_hybrid['alert_rate']:.2%}")
    synthetic_thresholds = pd.read_csv(ROOT / "models/synthetic_hybrid_thresholds.csv")
    add("Decision engine", "DEC-009", "Four validation-locked synthetic thresholds exist", len(synthetic_thresholds) == 4 and synthetic_thresholds["locked_threshold"].between(0, 1).all(), f"thresholds={len(synthetic_thresholds)}")

    alerts = pd.read_csv(ROOT / "outputs/alert_queue.csv.gz")
    add("Alert queue", "ALT-001", "Alert IDs unique", not alerts["alert_id"].duplicated().any(), f"alerts={len(alerts)}")
    add("Alert queue", "ALT-002", "All alerts have SLA and priority", alerts[["review_priority", "sla_hours"]].notna().all().all(), "priority/SLA")
    add("Alert queue", "ALT-003", "Handling time positive", alerts["handling_minutes"].gt(0).all(), f"min={alerts['handling_minutes'].min()}")
    add("Alert queue", "ALT-004", "Investigation outcomes complete", alerts["outcome"].notna().all(), "outcome")
    add("Alert queue", "ALT-005", "Capacity sensitivity includes breaches", pd.read_csv(ROOT / "outputs/operations_capacity_sensitivity.csv")["status"].eq("BREACH").any(), "stress scenario breach detectable")
    add("Alert queue", "ALT-006", "Investigation queue precision is at least 5%", alerts["transaction_fraud_label"].mean() >= 0.05, f"queue_precision={alerts['transaction_fraud_label'].mean():.2%}")
    add("Alert queue", "ALT-007", "Base average alert volume is within 250/day capacity", len(alerts) / 30 <= 250, f"alerts_per_day={len(alerts)/30:.1f}")

    incidents = pd.read_csv(ROOT / "operational_risk/incident_register.csv")
    add("Incident register", "INC-001", "At least 15 incidents", len(incidents) >= 15, f"incidents={len(incidents)}")
    add("Incident register", "INC-002", "Net loss reconciles", np.allclose(incidents["net_loss"], incidents["gross_loss"] - incidents["recovery"]), "net=gross-recovery")
    add("Incident register", "INC-003", "Loss values non-negative", incidents[["gross_loss", "recovery", "net_loss"]].ge(0).all().all(), "loss fields")
    add("Incident register", "INC-004", "Recovery not above gross loss", (incidents["recovery"] <= incidents["gross_loss"]).all(), "recovery<=gross")
    add("Incident register", "INC-005", "Every incident has owner/action/date/residual risk", incidents[["owner", "action_due_date", "status", "residual_risk"]].notna().all().all(), "incident governance fields")
    scenarios = pd.read_csv(ROOT / "operational_risk/scenario_analysis.csv")
    add("Incident register", "INC-006", "Ten required scenarios", len(scenarios) == 10, f"scenarios={len(scenarios)}")

    appetite = pd.read_csv(ROOT / "governance/fraud_risk_appetite_register.csv")
    add("Risk appetite", "APP-001", "Fraud appetite status populated", appetite["status"].isin(["GREEN", "AMBER", "RED"]).all(), str(appetite["status"].value_counts().to_dict()))
    add("Risk appetite", "APP-002", "Evidence layer declared per metric", appetite["evidence_layer"].notna().all(), "evidence_layer")
    add("Risk appetite", "APP-003", "Thresholds labelled educational", appetite["threshold_status"].str.contains("educational", case=False).all(), "threshold_status")
    fairness = pd.read_csv(ROOT / "validation/operational_segment_fairness.csv")
    fairness_actions = pd.read_csv(ROOT / "validation/operational_segment_fairness_actions.csv")
    flagged_dimensions = set(fairness.loc[fairness["review_trigger"].eq("FLAG"), "dimension"])
    action_dimensions = set(fairness_actions["dimension"].dropna())
    add("Risk appetite", "APP-004", "Every flagged operational disparity has an action register entry", flagged_dimensions.issubset(action_dimensions), f"flagged={sorted(flagged_dimensions)} actions={sorted(action_dimensions)}")
    add("Risk appetite", "APP-005", "Fairness actions include driver, mitigation, residual risk and owner", fairness_actions.empty or fairness_actions[["root_cause_driver", "mitigation", "residual_risk", "owner"]].notna().all().all(), f"action_rows={len(fairness_actions)}")

    for suite_name, relative, minimum in (
        ("UAT", "testing/uat_test_cases.csv", 30),
        ("SIT", "testing/sit_test_cases.csv", 15),
        ("Negative", "testing/negative_test_cases.csv", 15),
    ):
        suite = pd.read_csv(ROOT / relative)
        add("Testing", f"TEST-{suite_name}-COUNT", f"{suite_name} minimum case count", len(suite) >= minimum, f"count={len(suite)}")
        add("Testing", f"TEST-{suite_name}-PASS", f"All {suite_name} cases pass", suite["status"].eq("PASS").all(), str(suite["status"].value_counts().to_dict()))

    workbook_path = ROOT / "excel/Project5_FraudOperationalRisk_Model.xlsx"
    workbook = openpyxl.load_workbook(workbook_path, data_only=False, read_only=False)
    expected_sheets = [
        "00_Read_Me", "01_Observed_KPIs", "02_Model_Comparison", "03_Threshold_Economics",
        "04_Synthetic_Transactions", "05_Rule_Inventory", "06_Hybrid_Decisions", "07_Alert_Operations",
        "08_Control_Effectiveness", "09_Incident_Register", "10_KRI_Risk_Appetite", "11_UAT_SIT",
        "12_Reconciliation_Validation", "13_Executive_Summary",
    ]
    add("Excel", "XLS-001", "Workbook has exactly 14 governed sheets", workbook.sheetnames == expected_sheets, str(workbook.sheetnames))
    formulas = [cell.value for ws in workbook.worksheets for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    add("Excel", "XLS-002", "Workbook contains auditable formulas", len(formulas) >= 10, f"formula_count={len(formulas)}")
    add("Excel", "XLS-003", "No external workbook links", len(getattr(workbook, "_external_links", [])) == 0, f"external_links={len(getattr(workbook, '_external_links', []))}")
    add("Excel", "XLS-004", "Central assumptions present", workbook["00_Read_Me"]["B7"].value == 5.0 and workbook["00_Read_Me"]["B8"].value == 0.70, "Read Me central assumptions")

    recruiter_files = [ROOT / "README.md", ROOT / "recruiter_summary.md", ROOT / "limitation_note.md", ROOT / "OPEN_THIS_FIRST.html"]
    for index, path in enumerate(recruiter_files, start=1):
        content = path.read_text(encoding="utf-8")
        add("Claims", f"CLM-{index:02d}", f"Disclaimer/educational wording in {path.name}", "educational" in content.lower() or "not a production" in content.lower(), path.name)
        add("Claims", f"CLM-{index+10:02d}", f"Observed/synthetic separation in {path.name}", "observed" in content.lower() and "synthetic" in content.lower(), path.name)
    html_content = (ROOT / "OPEN_THIS_FIRST.html").read_text(encoding="utf-8")
    local_links = re.findall(r'href="([^"]+)"', html_content)
    relative_links = [link for link in local_links if not link.startswith(("http://", "https://", "#"))]
    unresolved_links = [link for link in relative_links if not (ROOT / link).exists()]
    add("Packaging", "PKG-001", "All local HTML links resolve", not unresolved_links, str(unresolved_links))
    add("Packaging", "PKG-002", "Requirements are pinned", all("==" in line for line in (ROOT / "requirements.txt").read_text().splitlines() if line.strip()), "requirements.txt")
    add("Packaging", "PKG-003", "No backslash href paths", '\\' not in "".join(local_links), "HTML href scan")
    add("Packaging", "PKG-004", "HTML has responsive viewport metadata", 'name="viewport"' in html_content, "meta viewport")
    add("Packaging", "PKG-005", "HTML has desktop/mobile breakpoints", "@media(max-width:850px)" in html_content and "@media(max-width:520px)" in html_content, "850px and 520px breakpoints")
    add("Packaging", "PKG-006", "All HTML tables have horizontal overflow wrappers", html_content.count("<table") == html_content.count('class="table-wrap"'), f"tables={html_content.count('<table')} wrappers={html_content.count('class=\"table-wrap\"')}")
    add("Packaging", "PKG-007", "No CSS gradient decoration", "gradient(" not in html_content.lower(), "CSS scan")
    add("Packaging", "PKG-008", "Letter spacing is non-negative", "letter-spacing:0" in html_content and "letter-spacing:-" not in html_content, "CSS scan")
    absolute_matches = absolute_path_matches()
    add("Packaging", "PKG-009", "No absolute personal Windows paths in public text/code files", not absolute_matches, str(absolute_matches))
    release = pd.read_csv(ROOT / "testing/release_readiness_checklist.csv")
    independent = pd.read_csv(ROOT / "validation/independent_review_checklist.csv")
    add("Governance claims", "GOV-001", "Independent review evidence is not labelled as real PASS/sign-off", independent["status"].eq("EVIDENCE_PREPARED").all(), str(independent["status"].value_counts().to_dict()))
    add("Governance claims", "GOV-002", "Independent organisational review and sign-off marked not performed", independent[["independent_review_performed", "independent_signoff_performed"]].eq("NO").all().all(), "independent flags")
    add("Governance claims", "GOV-003", "Release readiness is simulated, not organisational approval", release["status"].isin(["SIMULATED_PASS", "FAIL"]).all() and release["actual_organisational_approval"].eq("NO").all(), str(release["status"].value_counts().to_dict()))
    uncertainty = pd.read_csv(ROOT / "models/model_uncertainty_intervals.csv")
    add("Model metrics", "MOD-008", "Observed uncertainty intervals generated", {"fraud_recall", "precision", "pr_auc", "fraud_amount_recall"}.issubset(set(uncertainty["metric"])), str(sorted(uncertainty["metric"].unique())))
    priority_efficiency = pd.read_csv(ROOT / "outputs/priority_queue_efficiency.csv")
    add("Alert queue", "ALT-008", "Priority efficiency includes CRITICAL/HIGH/MEDIUM precision", {"CRITICAL", "HIGH", "MEDIUM"}.issubset(set(priority_efficiency["review_priority"])), str(priority_efficiency[["review_priority", "fraud_precision"]].to_dict(orient="records")))
    add("Source manifest", "SRC-001", "Observed source manifest present", (ROOT / "data_contract/source_manifest.csv").exists(), "source_manifest.csv")
    source_manifest = pd.read_csv(ROOT / "data_contract/source_manifest.csv")
    add("Source manifest", "SRC-002", "Source manifest has expanded public lineage columns", {"dataset_id", "source_url", "licence_or_access_terms", "canonical_path", "packaged_path", "downstream_use"}.issubset(source_manifest.columns), str(source_manifest.columns.tolist()))
    pre_manifest = build_artifact_manifest()
    add("Packaging", "PKG-010", "Artifact manifest coverage is at least 100 files and excludes itself", len(pre_manifest) >= 100 and "validation/artifact_manifest.csv" not in set(pre_manifest["relative_path"]), f"files={len(pre_manifest)}")

    result = pd.DataFrame(checks)
    result = add_metadata(result, "Mixed", "Derived", "Automated end-to-end validation", "Portfolio validation evidence")
    write_csv(result, "validation/validation_checks.csv")
    failed = result.loc[result["status"].eq("FAIL")]
    passed = int(result["status"].eq("PASS").sum())
    total = len(result)
    findings_table = dataframe_to_markdown(
        failed[["check_id", "check_group", "check_description", "evidence", "owner", "action_if_failed"]]
    ) if not failed.empty else "No unresolved findings."
    manifest = build_artifact_manifest()
    body = f"""## Final status

**{'PASS' if failed.empty else 'FAIL'} - {passed}/{total} checks passed.**

- Critical unresolved FAIL: **{len(failed.loc[failed['criticality'].eq('Critical')])}**.
- Total unresolved FAIL: **{len(failed)}**.
- Artifact manifest files: **{len(manifest)}**.
- Manifest is generated after validation outputs and excludes `validation/artifact_manifest.csv` to avoid self-referential hashing.
- Validation executed from: `{ROOT.name}` using relative project paths.

## Coverage

Files, observed data, synthetic data, label methodology, anti-circularity, models, calibration, threshold economics, rules, hybrid decisions, alert operations, capacity, incidents, risk appetite, UAT, SIT, negative tests, Excel, claims, links and source/artifact manifests were checked.

## Findings

{findings_table}

## Claim conclusion

Observed results remain confined to the PCA benchmark. Synthetic results remain confined to controls testing, alert operations, incidents and governance. The project does not claim a live payment platform, production fraud engine, AML system, legal fairness conclusion or regulatory operational-risk capital model.

Author automated self-validation completed. Independent model validation, production approval and organisational release sign-off were not performed.
"""
    write_markdown("validation/validation_report.md", "Project 5 Final Validation Report", body)

    audit_path = ROOT / "governance/audit_evidence_checklist.csv"
    audit = pd.read_csv(audit_path)
    audit.loc[audit["evidence_id"].eq("AE-10"), "status"] = "PASS" if failed.empty else "FAIL"
    write_csv(audit, "governance/audit_evidence_checklist.csv")

    final_log = ROOT / "logs/project5_v1_0_1_final_run.log"
    final_log.write_text(
        "\n".join(
            [
                "run_id=project5_v1_0_1_public_remediated",
                "project_version=Project5_FraudOperationalRisk_Gold_v1_0_1_PUBLIC_REMEDIATED",
                f"random_seed=20260711",
                f"package_mode=FULL_AND_AI_REVIEW_LIGHT",
                f"start_timestamp={start_ts}",
                f"end_timestamp={datetime.now().isoformat(timespec='seconds')}",
                "environment=local Python execution",
                f"validation_result={'PASS' if failed.empty else 'FAIL'}",
                f"validation_checks_passed={passed}",
                f"validation_checks_total={total}",
                f"artifact_manifest_files={len(manifest)}",
                "manifest_integrity_result=written_after_manifest",
                "independent_organisational_signoff=NO",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    manifest = build_artifact_manifest()
    write_csv(manifest, "validation/artifact_manifest.csv")
    integrity = write_manifest_integrity(manifest)
    print(f"FINAL VALIDATION {'PASS' if failed.empty else 'FAIL'} | {passed}/{total} checks | files={len(manifest)}")
    print(f"MANIFEST INTEGRITY {'PASS' if integrity['status'].eq('PASS').all() else 'FAIL'} | {int(integrity['status'].eq('PASS').sum())}/{len(integrity)} files")
    if not integrity["status"].eq("PASS").all():
        print(integrity.loc[~integrity["status"].eq("PASS"), ["relative_path", "exists", "hash_match", "duplicate_path"]].to_string(index=False))
        raise SystemExit(1)
    if not failed.empty:
        print(failed[["check_id", "check_description", "evidence"]].to_string(index=False))
        raise SystemExit(1)


if __name__ == "__main__":
    main()
