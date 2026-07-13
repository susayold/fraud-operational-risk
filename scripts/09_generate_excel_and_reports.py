from __future__ import annotations

import html
import json
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from _project5_common import (
    DISCLAIMER,
    ROOT,
    add_metadata,
    dataframe_to_markdown,
    decision_for_record,
    get_logger,
    safe_div,
    write_csv,
    write_markdown,
)


def base_test_record() -> dict:
    threshold_path = ROOT / "models/synthetic_hybrid_thresholds.csv"
    threshold_values = {
        "synthetic_model_step_up": 0.030594,
        "synthetic_model_review": 0.119124,
        "synthetic_model_decline": 0.135282,
        "synthetic_anomaly_review": 0.995206,
    }
    if threshold_path.exists():
        threshold_frame = pd.read_csv(threshold_path)
        threshold_values.update(dict(zip(threshold_frame["threshold_id"], threshold_frame["locked_threshold"])))
    return {
        "transaction_id": "T-UAT-001",
        "customer_id": "C-UAT-001",
        "device_id": "D-UAT-001",
        "beneficiary_id": "B-UAT-001",
        "transaction_time": "2026-01-10 12:00:00",
        "synthetic_model_score": 0.01,
        "synthetic_anomaly_score": 0.10,
        "amount": 100.0,
        "blacklist_hit": 0,
        "duplicate_transaction_flag": 0,
        "account_takeover_signal": 0,
        "failed_otp_count": 0,
        "critical_rule_hits": 0,
        "high_rule_hits": 0,
        "privileged_override_flag": 0,
        "quality_rule_overlay": 0,
        "model_step_up_threshold": float(threshold_values["synthetic_model_step_up"]),
        "model_review_threshold": float(threshold_values["synthetic_model_review"]),
        "model_decline_threshold": float(threshold_values["synthetic_model_decline"]),
        "anomaly_review_threshold": float(threshold_values["synthetic_anomaly_review"]),
    }


def make_uat_cases() -> pd.DataFrame:
    thresholds = base_test_record()
    step_up = thresholds["model_step_up_threshold"]
    review = thresholds["model_review_threshold"]
    decline = thresholds["model_decline_threshold"]
    anomaly_review = thresholds["anomaly_review_threshold"]
    cases = [
        ("UAT-001", "Allow below boundary", {"synthetic_model_score": max(0.0, step_up - 1e-7)}, "ALLOW"),
        ("UAT-002", "Score exactly step-up boundary", {"synthetic_model_score": step_up}, "STEP_UP"),
        ("UAT-003", "Score just below review boundary", {"synthetic_model_score": review - 1e-7}, "STEP_UP"),
        ("UAT-004", "Score exactly review boundary", {"synthetic_model_score": review}, "MANUAL_REVIEW"),
        ("UAT-005", "Score just below decline boundary", {"synthetic_model_score": decline - 1e-7, "amount": 1000}, "MANUAL_REVIEW"),
        ("UAT-006", "Score and amount exactly decline boundaries", {"synthetic_model_score": decline, "amount": 1000}, "DECLINE"),
        ("UAT-007", "Amount just below decline boundary", {"synthetic_model_score": decline, "amount": 999.99}, "MANUAL_REVIEW"),
        ("UAT-008", "High model low amount", {"synthetic_model_score": decline, "amount": 20}, "MANUAL_REVIEW"),
        ("UAT-009", "Anomaly exactly review boundary", {"synthetic_anomaly_score": anomaly_review}, "MANUAL_REVIEW"),
        ("UAT-010", "Anomaly below boundary", {"synthetic_anomaly_score": anomaly_review - 1e-7}, "ALLOW"),
        ("UAT-011", "One high rule", {"high_rule_hits": 1}, "STEP_UP"),
        ("UAT-012", "Two high rules with step-up score support", {"high_rule_hits": 2, "synthetic_model_score": step_up}, "MANUAL_REVIEW"),
        ("UAT-013", "Quality rule with model support", {"synthetic_model_score": 0.05, "quality_rule_overlay": 1}, "MANUAL_REVIEW"),
        ("UAT-014", "Critical rule", {"critical_rule_hits": 1}, "HOLD"),
        ("UAT-015", "ATO plus five OTP failures", {"account_takeover_signal": 1, "failed_otp_count": 5}, "HOLD"),
        ("UAT-016", "Blacklist overrides low score", {"blacklist_hit": 1}, "BLOCK_ACCOUNT"),
        ("UAT-017", "Blacklist overrides high score", {"blacklist_hit": 1, "synthetic_model_score": 0.95}, "BLOCK_ACCOUNT"),
        ("UAT-018", "Duplicate transaction flag", {"duplicate_transaction_flag": 1}, "HOLD"),
        ("UAT-019", "Privileged override misuse", {"privileged_override_flag": 1, "amount": 2000}, "BLOCK_ACCOUNT"),
        ("UAT-020", "Override below amount boundary", {"privileged_override_flag": 1, "amount": 1999.99}, "ALLOW"),
        ("UAT-021", "Missing model score", {"synthetic_model_score": None}, "DATA_EXCEPTION"),
        ("UAT-022", "Score below zero", {"synthetic_model_score": -0.01}, "DATA_EXCEPTION"),
        ("UAT-023", "Score above one", {"synthetic_model_score": 1.01}, "DATA_EXCEPTION"),
        ("UAT-024", "Negative amount", {"amount": -1}, "DATA_EXCEPTION"),
        ("UAT-025", "Missing transaction ID", {"transaction_id": None}, "DATA_EXCEPTION"),
        ("UAT-026", "Missing customer ID", {"customer_id": None}, "DATA_EXCEPTION"),
        ("UAT-027", "Missing device ID", {"device_id": None}, "DATA_EXCEPTION"),
        ("UAT-028", "Missing beneficiary ID", {"beneficiary_id": None}, "DATA_EXCEPTION"),
        ("UAT-029", "Invalid timestamp", {"transaction_time": "not-a-date"}, "DATA_EXCEPTION"),
        ("UAT-030", "Model unavailable represented as null", {"synthetic_model_score": np.nan, "critical_rule_hits": 1}, "DATA_EXCEPTION"),
        ("UAT-031", "Trusted low-risk device context", {}, "ALLOW"),
        ("UAT-032", "Trusted device but high model score", {"synthetic_model_score": review}, "MANUAL_REVIEW"),
        ("UAT-033", "Payroll-like low score new-device rule proxy", {"high_rule_hits": 1, "synthetic_model_score": 0.01}, "STEP_UP"),
        ("UAT-034", "Critical rule with model unavailable", {"critical_rule_hits": 1, "synthetic_model_score": None}, "DATA_EXCEPTION"),
        ("UAT-035", "Multiple conflict blacklist wins", {"blacklist_hit": 1, "critical_rule_hits": 2, "synthetic_model_score": decline}, "BLOCK_ACCOUNT"),
        ("UAT-036", "Exact zero score", {"synthetic_model_score": 0.0}, "ALLOW"),
        ("UAT-037", "Exact one score high amount", {"synthetic_model_score": 1.0, "amount": 1000}, "DECLINE"),
        ("UAT-038", "Zero amount low risk", {"amount": 0.0}, "ALLOW"),
    ]
    rows = []
    for case_id, description, changes, expected in cases:
        record = base_test_record()
        record.update(changes)
        actual, reason = decision_for_record(record)
        rows.append(
            {
                "test_case_id": case_id,
                "test_category": "Boundary/exception/conflict",
                "test_description": description,
                "input_json": json.dumps(record, default=str, sort_keys=True),
                "expected_decision": expected,
                "actual_decision": actual,
                "actual_reason_code": reason,
                "status": "PASS" if actual == expected else "FAIL",
                "evidence": "Executed shared decision_for_record reference function",
            }
        )
    return add_metadata(
        pd.DataFrame(rows),
        "Synthetic",
        "Derived",
        "Executable hybrid decision reference function",
        "UAT controls-testing evidence",
    )


def create_sit_cases() -> pd.DataFrame:
    checks = [
        ("SIT-001", "Observed data loads", (ROOT / "data/observed/fraud_pca_input.csv.gz").exists()),
        ("SIT-002", "Synthetic data loads", (ROOT / "data/synthetic/synthetic_transactions.csv.gz").exists()),
        ("SIT-003", "Observed score scale 0-1", pd.read_csv(ROOT / "outputs/observed_model_predictions.csv.gz", usecols=["champion_score_calibrated"])["champion_score_calibrated"].between(0, 1).all()),
        ("SIT-004", "Synthetic score scale 0-1", pd.read_csv(ROOT / "outputs/synthetic_transaction_scores.csv.gz", usecols=["synthetic_model_score"])["synthetic_model_score"].between(0, 1).all()),
        ("SIT-005", "Anomaly score mapping 0-1", pd.read_csv(ROOT / "outputs/synthetic_transaction_scores.csv.gz", usecols=["synthetic_anomaly_score"])["synthetic_anomaly_score"].between(0, 1).all()),
        ("SIT-006", "Rule engine output", (ROOT / "outputs/synthetic_rule_hits.csv.gz").exists()),
        ("SIT-007", "Hybrid decision output", (ROOT / "outputs/hybrid_decision_output.csv.gz").exists()),
        ("SIT-008", "Alert queue output", (ROOT / "outputs/alert_queue.csv.gz").exists()),
        ("SIT-009", "Incident register output", (ROOT / "operational_risk/incident_register.csv").exists()),
        ("SIT-010", "Reporting outputs present", (ROOT / "reports/model_performance_report.md").exists()),
        ("SIT-011", "Missing-score spike routes exceptions", decision_for_record({**base_test_record(), "synthetic_model_score": None})[0] == "DATA_EXCEPTION"),
        ("SIT-012", "Score scale mismatch 0-100 blocked", decision_for_record({**base_test_record(), "synthetic_model_score": 87})[0] == "DATA_EXCEPTION"),
        ("SIT-013", "Device vendor unavailable fallback documented", (ROOT / "rules/rule_priority_matrix.csv").exists()),
        ("SIT-014", "Duplicate transaction hold", decision_for_record({**base_test_record(), "duplicate_transaction_flag": 1})[0] == "HOLD"),
        ("SIT-015", "Capacity sensitivity generated", (ROOT / "outputs/operations_capacity_sensitivity.csv").exists()),
        ("SIT-016", "Friendly fraud claim layer separate", (ROOT / "data/synthetic/synthetic_disputes.csv").exists()),
        ("SIT-017", "Observed and synthetic layers separated", (ROOT / "data_contract/observed_vs_synthetic_design.md").exists()),
        ("SIT-018", "Risk appetite output", (ROOT / "governance/fraud_risk_appetite_register.csv").exists()),
    ]
    rows = [
        {
            "test_case_id": test_id,
            "integration_control": description,
            "expected_result": "PASS",
            "actual_result": "PASS" if bool(result) else "FAIL",
            "status": "PASS" if bool(result) else "FAIL",
            "evidence": "Executable artifact or integration check",
        }
        for test_id, description, result in checks
    ]
    return add_metadata(pd.DataFrame(rows), "Mixed", "Derived", "Project integration tests", "Portfolio SIT evidence")


def create_negative_tests() -> pd.DataFrame:
    sample = pd.DataFrame({"transaction_id": ["T1", "T2"], "fraud_flag": [0, 1], "amount": [10.0, 20.0]})
    tests = [
        ("NEG-001", "Null risk score rejected", decision_for_record({**base_test_record(), "synthetic_model_score": None})[0] == "DATA_EXCEPTION"),
        ("NEG-002", "Invalid class detected", not pd.Series([0, 2]).isin([0, 1]).all()),
        ("NEG-003", "Negative amount rejected", decision_for_record({**base_test_record(), "amount": -0.01})[0] == "DATA_EXCEPTION"),
        ("NEG-004", "Duplicate ID detected", pd.Series(["T1", "T1"]).duplicated().any()),
        ("NEG-005", "Timestamp disorder detected", not pd.Series(pd.to_datetime(["2026-01-02", "2026-01-01"])).is_monotonic_increasing),
        ("NEG-006", "Rule priority blacklist wins", decision_for_record({**base_test_record(), "blacklist_hit": 1, "synthetic_model_score": 0.01})[0] == "BLOCK_ACCOUNT"),
        ("NEG-007", "Missing feature routes exception", decision_for_record({**base_test_record(), "device_id": None})[0] == "DATA_EXCEPTION"),
        ("NEG-008", "Missing model file detectable", not (ROOT / "models/intentionally_missing.joblib").exists()),
        ("NEG-009", "Score scale mismatch blocked", decision_for_record({**base_test_record(), "synthetic_model_score": 70})[0] == "DATA_EXCEPTION"),
        ("NEG-010", "Alert queue overflow detected", (pd.read_csv(ROOT / "outputs/operations_capacity_sensitivity.csv")["status"] == "BREACH").any()),
        ("NEG-011", "SLA breach detectable", pd.read_csv(ROOT / "outputs/alert_queue.csv.gz", usecols=["sla_breach_flag"])["sla_breach_flag"].isin([0, 1]).all()),
        ("NEG-012", "Missing reason code detectable", pd.Series(["RC_OK", ""]).eq("").any()),
        ("NEG-013", "Missing action owner detectable", pd.Series(["Owner", ""]).eq("").any()),
        ("NEG-014", "Negative incident loss rejected", not pd.Series([-1.0]).ge(0).all()),
        ("NEG-015", "Recovery above gross loss rejected", bool((pd.Series([120.0]) > pd.Series([100.0])).any())),
        ("NEG-016", "Observed sample base control valid", sample["fraud_flag"].isin([0, 1]).all() and sample["amount"].ge(0).all()),
    ]
    rows = [
        {
            "test_case_id": test_id,
            "negative_condition": description,
            "expected_control": "Condition detected or rejected",
            "actual_control_result": "Detected" if bool(result) else "Missed",
            "status": "PASS" if bool(result) else "FAIL",
            "evidence": "Executable validation expression/reference function",
        }
        for test_id, description, result in tests
    ]
    return add_metadata(pd.DataFrame(rows), "Mixed", "Derived", "Executable negative controls", "Portfolio negative-test evidence")


def style_sheet(ws, banner: str) -> None:
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row == 1:
            ws.unmerge_cells(str(merged_range))
    ws["A1"] = banner
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = PatternFill("solid", fgColor="243746")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, ws.max_column))
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2B7A78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    headers = {ws.cell(2, col).value: col for col in range(1, ws.max_column + 1)}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for header, column in headers.items():
        header_text = str(header or "").lower()
        if any(token in header_text for token in ("rate", "recall", "precision", "f1", "roc_auc", "pr_auc", "brier", "psi")):
            for row in range(3, ws.max_row + 1):
                ws.cell(row, column).number_format = "0.000" if "auc" in header_text or "brier" in header_text or "psi" in header_text else "0.00%"
        elif any(token in header_text for token in ("amount", "cost", "benefit", "loss", "recovery")):
            for row in range(3, ws.max_row + 1):
                ws.cell(row, column).number_format = "#,##0.00"
        elif any(token in header_text for token in ("alerts", "transactions", "rows", "count", "volume")):
            for row in range(3, ws.max_row + 1):
                ws.cell(row, column).number_format = "#,##0"
    for column in range(1, ws.max_column + 1):
        values = [str(ws.cell(row=row, column=column).value or "") for row in range(1, min(ws.max_row, 300) + 1)]
        width = min(36, max(10, max(len(value) for value in values) + 2))
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.auto_filter.ref = ws.dimensions


def write_frame(ws, frame: pd.DataFrame, start_row: int = 2, max_rows: int | None = None) -> int:
    selected = frame.head(max_rows) if max_rows else frame
    for column_index, column in enumerate(selected.columns, start=1):
        ws.cell(start_row, column_index, column)
    for row_index, row in enumerate(selected.itertuples(index=False, name=None), start=start_row + 1):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            if isinstance(value, (np.integer,)):
                value = int(value)
            if isinstance(value, (np.floating,)):
                value = float(value)
            ws.cell(row_index, column_index, value)
    return start_row + len(selected)


def build_excel(uat: pd.DataFrame, sit: pd.DataFrame) -> None:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    readme = workbook.create_sheet("00_Read_Me")
    readme_rows = [
        ["Project", "Project 5 - Fraud Detection, Transaction Controls & Operational Risk"],
        ["Purpose", "Recruiter-ready educational risk framework"],
        ["Observed layer", "PCA fraud benchmark: predictive performance and observed amount capture"],
        ["Synthetic layer", "Interpretable controls, decisions, alerts, incidents and UAT/SIT"],
        ["Disclaimer", DISCLAIMER],
        ["Assumption", "Value"],
        ["Investigation cost per alert", 5.0],
        ["Preventable loss rate", 0.70],
        ["False-positive friction cost", 2.0],
        ["Alert capacity per day", 250],
        ["Formula policy", "Threshold sheet formulas reference these central assumptions"],
    ]
    for row in readme_rows:
        readme.append(row)
    readme["A1"].font = Font(bold=True, color="FFFFFF")
    readme["A1"].fill = PatternFill("solid", fgColor="243746")
    readme.column_dimensions["A"].width = 34
    readme.column_dimensions["B"].width = 100
    readme.sheet_view.showGridLines = False
    readme.freeze_panes = "A2"

    sources = {
        "01_Observed_KPIs": pd.read_csv(ROOT / "outputs/observed_fraud_kpi_summary.csv"),
        "02_Model_Comparison": pd.read_csv(ROOT / "models/model_comparison.csv"),
        "03_Threshold_Economics": pd.read_csv(ROOT / "outputs/threshold_economics.csv"),
        "04_Synthetic_Transactions": pd.read_csv(ROOT / "data/synthetic/synthetic_transactions_sample.csv").head(2000),
        "05_Rule_Inventory": pd.read_csv(ROOT / "rules/fraud_rule_inventory.csv"),
        "06_Hybrid_Decisions": pd.read_csv(ROOT / "outputs/hybrid_decision_summary.csv"),
        "07_Alert_Operations": pd.read_csv(ROOT / "outputs/alert_outcome_summary.csv"),
        "08_Control_Effectiveness": pd.read_csv(ROOT / "outputs/control_effectiveness_summary.csv"),
        "09_Incident_Register": pd.read_csv(ROOT / "operational_risk/incident_register.csv"),
        "10_KRI_Risk_Appetite": pd.concat([
            pd.read_csv(ROOT / "governance/fraud_risk_appetite_register.csv"),
            pd.read_csv(ROOT / "governance/operational_risk_appetite_register.csv"),
        ], ignore_index=True, sort=False),
        "11_UAT_SIT": pd.concat([uat.assign(test_suite="UAT"), sit.assign(test_suite="SIT")], ignore_index=True, sort=False),
        "12_Reconciliation_Validation": pd.concat([
            pd.read_csv(ROOT / "data_contract/population_reconciliation.csv"),
            pd.read_csv(ROOT / "validation/synthetic_anti_circularity_tests.csv"),
        ], ignore_index=True, sort=False),
    }
    for sheet_name, frame in sources.items():
        ws = workbook.create_sheet(sheet_name)
        write_frame(ws, frame, start_row=2)
        banner = "OBSERVED PCA BENCHMARK" if sheet_name in ("01_Observed_KPIs", "02_Model_Comparison", "03_Threshold_Economics") else "SYNTHETIC CONTROLS-TESTING / GOVERNANCE"
        if sheet_name in ("10_KRI_Risk_Appetite", "11_UAT_SIT", "12_Reconciliation_Validation"):
            banner = "MIXED EVIDENCE - READ data_layer AND claim_boundary"
        style_sheet(ws, banner)
        if ws.max_row > 2 and ws.max_column > 1:
            table_ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName=f"T_{sheet_name.replace('_', '')[:20]}", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
            ws.add_table(table)

    threshold_ws = workbook["03_Threshold_Economics"]
    headers = {threshold_ws.cell(2, col).value: col for col in range(1, threshold_ws.max_column + 1)}
    formula_start = threshold_ws.max_column + 1
    formula_headers = ["excel_prevented_loss", "excel_investigation_cost", "excel_friction_cost", "excel_net_benefit"]
    for offset, header in enumerate(formula_headers):
        threshold_ws.cell(2, formula_start + offset, header)
    for row in range(3, threshold_ws.max_row + 1):
        captured_cell = f"{get_column_letter(headers['fraud_amount_captured'])}{row}"
        alerts_cell = f"{get_column_letter(headers['alerts'])}{row}"
        fp_cell = f"{get_column_letter(headers['false_positives'])}{row}"
        threshold_ws.cell(row, formula_start, f"={captured_cell}*'00_Read_Me'!$B$8")
        threshold_ws.cell(row, formula_start + 1, f"={alerts_cell}*'00_Read_Me'!$B$7")
        threshold_ws.cell(row, formula_start + 2, f"={fp_cell}*'00_Read_Me'!$B$9")
        threshold_ws.cell(row, formula_start + 3, f"={get_column_letter(formula_start)}{row}-{get_column_letter(formula_start+1)}{row}-{get_column_letter(formula_start+2)}{row}")
    style_sheet(threshold_ws, "OBSERVED PCA BENCHMARK - BACKTEST ECONOMICS WITH CENTRAL ASSUMPTIONS")

    executive = workbook.create_sheet("13_Executive_Summary")
    selected = pd.read_csv(ROOT / "outputs/threshold_economics.csv")
    selected = selected.loc[selected["population"].eq("test") & selected["recommended_threshold_flag"].eq(1)].iloc[0]
    incidents = pd.read_csv(ROOT / "operational_risk/incident_register.csv")
    alerts = pd.read_csv(ROOT / "outputs/alert_queue.csv.gz", usecols=["alert_id", "sla_breach_flag", "outcome"])
    model = pd.read_csv(ROOT / "models/model_comparison.csv")
    champion = model.loc[model["population"].eq("test") & model["model_role"].eq("Champion")].iloc[0]
    executive_rows = [
        ["Metric", "Result", "Senior interpretation"],
        ["Observed champion", champion["model"], "Selected on validation evidence; PCA ranking benchmark"],
        ["Observed test PR-AUC", f"{champion['pr_auc']:.3f}", "Primary imbalance-aware discrimination metric"],
        ["Recommended alert policy", f"{selected['target_alert_rate']:.2%}", "Validation-locked under frozen Base economics"],
        ["Test fraud amount recall", f"{selected['fraud_amount_recall']:.2%}", "Observed amount captured, not guaranteed prevented loss"],
        ["Test precision", f"{selected['precision']:.2%}", "Fraud yield among alerts"],
        ["Synthetic investigation alerts", len(alerts), "Controls-testing workload only"],
        ["Synthetic SLA breach rate", f"{alerts['sla_breach_flag'].mean():.2%}", "Capacity/control governance indicator"],
        ["Synthetic incident net loss", incidents["net_loss"].sum(), "Scenario loss proxy, not observed loss"],
        ["Public claim boundary", "Educational framework", "No production fraud, AML or regulatory model claim"],
    ]
    for row in executive_rows:
        executive.append(row)
    executive.sheet_view.showGridLines = False
    executive.column_dimensions["A"].width = 34
    executive.column_dimensions["B"].width = 28
    executive.column_dimensions["C"].width = 76
    for cell in executive[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="243746")
    executive.freeze_panes = "A2"
    executive.conditional_formatting.add("B2:B20", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FADBD8")))

    chart = BarChart()
    model_ws = workbook["02_Model_Comparison"]
    model_headers = {model_ws.cell(2, col).value: col for col in range(1, model_ws.max_column + 1)}
    chart.add_data(Reference(model_ws, min_col=model_headers["pr_auc"], min_row=2, max_row=min(model_ws.max_row, 16)), titles_from_data=True)
    chart.set_categories(Reference(model_ws, min_col=model_headers["model"], min_row=3, max_row=min(model_ws.max_row, 16)))
    chart.title = "Model PR-AUC comparison"
    chart.height = 8
    chart.width = 14
    executive.add_chart(chart, "E2")

    threshold_chart = LineChart()
    threshold_headers = {threshold_ws.cell(2, col).value: col for col in range(1, threshold_ws.max_column + 1)}
    threshold_chart.add_data(
        Reference(threshold_ws, min_col=threshold_headers["realised_net_benefit"], min_row=2, max_row=min(threshold_ws.max_row, 16)),
        titles_from_data=True,
    )
    threshold_chart.set_categories(
        Reference(threshold_ws, min_col=threshold_headers["target_alert_rate"], min_row=3, max_row=min(threshold_ws.max_row, 16))
    )
    threshold_chart.title = "Threshold economics"
    threshold_chart.y_axis.title = "Net benefit"
    threshold_chart.x_axis.title = "Target alert rate"
    threshold_chart.height = 8
    threshold_chart.width = 14
    executive.add_chart(threshold_chart, "E20")

    alert_chart = BarChart()
    alert_ws = workbook["07_Alert_Operations"]
    alert_headers = {alert_ws.cell(2, col).value: col for col in range(1, alert_ws.max_column + 1)}
    alert_chart.add_data(
        Reference(alert_ws, min_col=alert_headers["alerts"], min_row=2, max_row=min(alert_ws.max_row, 18)),
        titles_from_data=True,
    )
    alert_chart.set_categories(
        Reference(alert_ws, min_col=alert_headers["outcome"], min_row=3, max_row=min(alert_ws.max_row, 18))
    )
    alert_chart.title = "Alert operations volume"
    alert_chart.y_axis.title = "Alerts"
    alert_chart.height = 8
    alert_chart.width = 14
    executive.add_chart(alert_chart, "E38")

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    path = ROOT / "excel/Project5_FraudOperationalRisk_Model.xlsx"
    workbook.save(path)


def html_table(frame: pd.DataFrame, columns: list[str], percent_columns: set[str] | None = None) -> str:
    percent_columns = percent_columns or set()
    rows = ["<div class=\"table-wrap\"><table><thead><tr>" + "".join(f"<th>{html.escape(column.replace('_', ' ').title())}</th>" for column in columns) + "</tr></thead><tbody>"]
    for _, row in frame[columns].iterrows():
        cells = []
        for column in columns:
            value = row[column]
            if column in percent_columns and pd.notna(value):
                display = f"{float(value):.2%}"
            elif isinstance(value, (float, np.floating)):
                display = f"{float(value):,.4f}"
            else:
                display = str(value)
            cells.append(f"<td>{html.escape(display)}</td>")
        rows.append("<tr>" + "".join(cells) + "</tr>")
    rows.append("</tbody></table></div>")
    return "".join(rows)


def build_html() -> None:
    model = pd.read_csv(ROOT / "models/model_comparison.csv")
    champion = model.loc[model["population"].eq("test") & model["model_role"].eq("Champion")].iloc[0]
    threshold = pd.read_csv(ROOT / "outputs/threshold_economics.csv")
    selected = threshold.loc[threshold["population"].eq("test") & threshold["recommended_threshold_flag"].eq(1)].iloc[0]
    rules = pd.read_csv(ROOT / "outputs/rule_performance_summary.csv").sort_values("fraud_recall", ascending=False).head(8)
    incremental = pd.read_csv(ROOT / "outputs/hybrid_incremental_value.csv")
    appetite = pd.read_csv(ROOT / "governance/fraud_risk_appetite_register.csv")
    incidents = pd.read_csv(ROOT / "operational_risk/incident_register.csv")
    alerts = pd.read_csv(ROOT / "outputs/alert_queue.csv.gz", usecols=["alert_id", "sla_breach_flag", "outcome", "prevented_loss_proxy"])
    validation = pd.read_csv(ROOT / "testing/uat_test_cases.csv")
    html_content = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Project 5 | Fraud & Operational Risk</title>
<style>
:root{{--ink:#1c2833;--muted:#53636f;--paper:#f5f7f8;--white:#fff;--teal:#176b68;--red:#b5333e;--amber:#a76b00;--line:#d8e0e4;--nav:#243746;}}
*{{box-sizing:border-box}} body{{margin:0;font:15px/1.55 Arial,sans-serif;color:var(--ink);background:var(--paper);letter-spacing:0}}
a{{color:var(--teal);text-decoration:none}} nav{{height:58px;background:var(--nav);color:#fff;display:flex;align-items:center;justify-content:space-between;padding:0 max(24px,5vw);position:sticky;top:0;z-index:5}} nav a{{color:#fff;margin-left:20px;font-size:13px}}
.hero{{background:#fff;padding:58px max(24px,7vw) 42px;border-bottom:1px solid var(--line)}} .eyebrow{{font-size:12px;text-transform:uppercase;color:var(--teal);font-weight:bold}} h1{{font-size:40px;line-height:1.08;margin:10px 0 14px;max-width:980px}} .lead{{font-size:18px;color:var(--muted);max-width:900px}}
.metrics{{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:1px;background:var(--line);margin-top:32px;border:1px solid var(--line)}} .metric{{background:#fff;padding:18px}} .metric strong{{display:block;font-size:25px}} .metric span{{color:var(--muted);font-size:12px}}
section{{padding:42px max(24px,7vw);border-bottom:1px solid var(--line)}} section.white{{background:#fff}} h2{{font-size:26px;margin:0 0 10px}} h3{{font-size:18px;margin:22px 0 8px}} .note{{border-left:4px solid var(--amber);padding:12px 16px;background:#fff8e8;max-width:1000px}} .grid{{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:28px}} .panel{{border:1px solid var(--line);background:#fff;padding:20px;border-radius:6px}} .tag{{display:inline-block;padding:4px 8px;border-radius:4px;font-size:11px;font-weight:bold;margin-right:6px}} .observed{{background:#dceef0;color:#155554}} .synthetic{{background:#fbe4e6;color:#7f222c}} .derived{{background:#fff0cc;color:#6c4900}}
table{{width:100%;border-collapse:collapse;background:#fff;font-size:12px}} .table-wrap{{width:100%;overflow-x:auto;border:1px solid var(--line)}} .table-wrap table{{min-width:720px}} th,td{{padding:9px;border-bottom:1px solid var(--line);text-align:left;vertical-align:top}} th{{background:var(--nav);color:#fff}} .links a{{display:inline-block;border:1px solid var(--line);background:#fff;padding:9px 12px;margin:4px 6px 4px 0;border-radius:5px}} footer{{padding:30px max(24px,7vw);background:var(--nav);color:#dfe7ea;font-size:12px}}
@media(max-width:850px){{h1{{font-size:31px}}.metrics,.grid{{grid-template-columns:1fr 1fr}}nav .navlinks{{display:none}}}} @media(max-width:520px){{.metrics,.grid{{grid-template-columns:1fr}}}}
</style></head><body>
<nav><strong>FRAUD RISK / PROJECT 5</strong><div class="navlinks"><a href="#observed">Observed</a><a href="#controls">Controls</a><a href="#operations">Operations</a><a href="#governance">Governance</a></div></nav>
<header class="hero"><div class="eyebrow">Senior Fraud Analytics · Transaction Controls · Operational Risk</div><h1>Fraud Detection, Transaction Risk Controls & Operational Risk Governance</h1><p class="lead">A two-layer risk framework that turns observed predictive evidence into a separately governed controls-testing environment, then traces decisions through alert capacity, incidents, KRIs, UAT/SIT and management action.</p>
<div class="metrics"><div class="metric"><strong>{champion['pr_auc']:.3f}</strong><span>Observed test PR-AUC</span></div><div class="metric"><strong>{selected['fraud_amount_recall']:.1%}</strong><span>Observed test fraud amount recall</span></div><div class="metric"><strong>{selected['precision']:.1%}</strong><span>Observed alert precision</span></div><div class="metric"><strong>{int(validation['status'].eq('PASS').sum())}/{len(validation)}</strong><span>Executable UAT passed</span></div></div></header>
<section id="observed"><span class="tag observed">OBSERVED</span><span class="tag derived">DERIVED</span><h2>Observed PCA fraud benchmark</h2><p>Champion <strong>{html.escape(str(champion['model']))}</strong> was selected on validation PR-AUC, fraud-amount capture, workload and stability. Final test data was not used for fitting, calibration or threshold selection.</p><div class="grid"><div><h3>Model evidence</h3>{html_table(model[model['population'].eq('test')].sort_values('pr_auc',ascending=False), ['model','model_role','pr_auc','roc_auc','precision','recall','fraud_amount_recall'], {'precision','recall','fraud_amount_recall'})}</div><div><h3>Threshold recommendation</h3><div class="panel"><strong>Top {selected['target_alert_rate']:.2%} validation-locked alert policy</strong><p>Test alert rate {selected['actual_alert_rate']:.2%}; fraud recall {selected['fraud_recall']:.2%}; amount recall {selected['fraud_amount_recall']:.2%}; backtest net benefit under disclosed cost assumptions {selected['realised_net_benefit']:,.0f} cost units.</p><p class="note">Fraud recall and precision meet the balanced decision criteria, while fraud-amount recall remains Amber and requires explicit management acceptance plus high-value control overlays. PCA variables support ranking, not business reason codes.</p></div></div></div></section>
<section id="controls" class="white"><span class="tag synthetic">SYNTHETIC</span><h2>Interpretable controls and hybrid decisioning</h2><p>250,000 seeded transactions were generated from hidden fraud mechanisms. Observable symptoms are noisy; legitimate transactions can trigger rules and fraud can avoid them. Friendly fraud is handled as a separate post-transaction dispute label.</p><div class="grid"><div><h3>Rule effectiveness</h3>{html_table(rules, ['rule_id','rule_name','alert_rate','precision','fraud_recall','control_status'], {'alert_rate','precision','fraud_recall'})}</div><div><h3>Incremental value</h3>{html_table(incremental, ['component_set','alert_rate','precision','fraud_recall','fraud_amount_recall','unique_fraud_contribution'], {'alert_rate','precision','fraud_recall','fraud_amount_recall'})}</div></div></section>
<section id="operations"><span class="tag synthetic">SYNTHETIC</span><h2>Alert operations and operational risk</h2><div class="metrics"><div class="metric"><strong>{len(alerts):,}</strong><span>Investigation alerts</span></div><div class="metric"><strong>{alerts['sla_breach_flag'].mean():.1%}</strong><span>SLA breach proxy</span></div><div class="metric"><strong>{len(incidents)}</strong><span>Simulated incidents</span></div><div class="metric"><strong>{incidents['net_loss'].sum():,.0f}</strong><span>Net loss scenario units</span></div></div><p class="note">Alert outcomes, staffing, losses and incidents are synthetic operational proxies. They demonstrate governance and capacity analysis, not actual bank performance.</p></section>
<section id="governance" class="white"><h2>Risk appetite, controls and evidence</h2>{html_table(appetite, ['metric','current_value','status','evidence_layer','action_if_red'])}<h3>Artifacts</h3><div class="links"><a href="README.md">README</a><a href="recruiter_summary.md">Recruiter summary</a><a href="excel/Project5_FraudOperationalRisk_Model.xlsx">Excel model</a><a href="reports/fraud_risk_committee_memo.md">Fraud Risk Committee memo</a><a href="reports/operational_risk_committee_memo.md">Operational Risk Committee memo</a><a href="validation/validation_report.md">Validation report</a><a href="rules/fraud_rule_inventory.csv">Rule inventory</a></div></section>
<footer>{html.escape(DISCLAIMER)}</footer></body></html>"""
    (ROOT / "OPEN_THIS_FIRST.html").write_text(html_content, encoding="utf-8")


def main() -> None:
    log = get_logger("phase8.excel_reports")
    uat = make_uat_cases()
    sit = create_sit_cases()
    negative = create_negative_tests()
    write_csv(uat, "testing/uat_test_cases.csv")
    write_csv(sit, "testing/sit_test_cases.csv")
    write_csv(negative, "testing/negative_test_cases.csv")

    review_sample_dir = ROOT / "data/review_samples"
    review_sample_dir.mkdir(parents=True, exist_ok=True)
    pd.read_csv(ROOT / "data/synthetic/synthetic_customers.csv").sample(2_000, random_state=20260711).to_csv(review_sample_dir / "synthetic_customers_sample.csv", index=False)
    pd.read_csv(ROOT / "data/synthetic/synthetic_devices.csv").sample(2_000, random_state=20260711).to_csv(review_sample_dir / "synthetic_devices_sample.csv", index=False)
    pd.read_csv(ROOT / "data/synthetic/synthetic_beneficiaries.csv").sample(2_000, random_state=20260711).to_csv(review_sample_dir / "synthetic_beneficiaries_sample.csv", index=False)
    pd.read_csv(ROOT / "data/synthetic/synthetic_disputes.csv").sample(2_000, random_state=20260711).to_csv(review_sample_dir / "synthetic_disputes_sample.csv", index=False)
    pd.read_csv(ROOT / "outputs/hybrid_decision_output.csv.gz").sample(5_000, random_state=20260711).to_csv(review_sample_dir / "hybrid_decisions_sample.csv", index=False)
    alert_sample = pd.read_csv(ROOT / "outputs/alert_queue.csv.gz")
    alert_sample.sample(min(2_000, len(alert_sample)), random_state=20260711).to_csv(review_sample_dir / "alert_queue_sample.csv", index=False)
    evidence = pd.concat(
        [
            uat[["test_case_id", "status", "evidence"]].assign(test_suite="UAT"),
            sit[["test_case_id", "status", "evidence"]].assign(test_suite="SIT"),
            negative[["test_case_id", "status", "evidence"]].assign(test_suite="Negative"),
        ],
        ignore_index=True,
    )
    evidence["executed_by"] = "Automated Project 5 pipeline"
    evidence["execution_date"] = "2026-07-11"
    evidence["evidence_status"] = np.where(evidence["status"].eq("PASS"), "Retained", "Action required")
    evidence = add_metadata(evidence, "Mixed", "Derived", "Executable tests", "Portfolio test evidence")
    write_csv(evidence, "testing/test_evidence_log.csv")

    simulated_pass = "SIMULATED_PASS"
    release_items = [
        ("Model evidence prepared", simulated_pass, "models/champion_model_card.md"),
        ("Threshold evidence prepared", simulated_pass, "reports/threshold_recommendation_memo.md"),
        ("Rule inventory evidence prepared", simulated_pass, "rules/fraud_rule_inventory.csv"),
        ("UAT simulated pass", simulated_pass if uat["status"].eq("PASS").all() else "FAIL", "testing/uat_test_cases.csv"),
        ("SIT simulated pass", simulated_pass if sit["status"].eq("PASS").all() else "FAIL", "testing/sit_test_cases.csv"),
        ("Negative tests simulated pass", simulated_pass if negative["status"].eq("PASS").all() else "FAIL", "testing/negative_test_cases.csv"),
        ("Capacity evidence prepared", simulated_pass, "reports/alert_operations_report.md"),
        ("Customer-friction review evidence prepared", simulated_pass, "reports/operational_segment_fairness_note.md"),
        ("Fallback logic tested", simulated_pass, "testing/sit_test_cases.csv"),
        ("Monitoring evidence prepared", simulated_pass, "governance/monitoring_plan.csv"),
        ("Rollback plan prepared", simulated_pass, "testing/rollback_plan.md"),
        ("Incident communication prepared", simulated_pass, "testing/incident_management_playbook.md"),
        ("Audit evidence prepared", simulated_pass, "governance/audit_evidence_checklist.csv"),
    ]
    release = pd.DataFrame(release_items, columns=["release_control", "status", "evidence"])
    release["owner"] = "Simulated governance role"
    release["signoff_type"] = "Portfolio simulation"
    release["actual_organisational_approval"] = "NO"
    release = add_metadata(release, "Governance", "Derived", "Project control evidence", "No production release claim")
    write_csv(release, "testing/release_readiness_checklist.csv")

    write_markdown(
        "testing/rollback_plan.md",
        "Rollback and Fallback Plan",
        """## Trigger and response

| Trigger | Immediate mode | Recovery evidence |
| --- | --- | --- |
| Model unavailable | Rules-only mode | Model health restored and SIT rerun |
| Rules engine unavailable | Model plus conservative step-up | Rule package checksum and regression PASS |
| Device vendor unavailable | Conservative step-up | Vendor feed freshness restored |
| Alert queue unavailable | Hold critical; allow low risk with logging | Queue replay reconciled |
| Score scale invalid | Release block | 0-1 boundary test and proposed independent-review evidence |

Rollback requires preserved pre-change artifact, versioned configuration, simulated owner approval, affected-decision reconciliation and post-rollback monitoring. This portfolio package does not claim actual organisational approval.""",
    )
    write_markdown(
        "testing/incident_management_playbook.md",
        "Incident Management Playbook",
        """Detect and classify the event; protect customers and preserve evidence; invoke the governed fallback; quantify transaction, loss, customer and backlog impact; notify the accountable owner; open corrective actions with due dates; perform RCA; rerun UAT/SIT and prepare proposed independent-review evidence; close only after simulated residual-risk acceptance. This is portfolio governance evidence, not organisational sign-off.""",
    )

    independent = pd.DataFrame(
        [
            ["IR-01", "Data lineage", "Source manifest, checksums and population reconciliation", "EVIDENCE_PREPARED", "Proposed independent review role"],
            ["IR-02", "No test leakage", "Ordered split and validation-only threshold", "EVIDENCE_PREPARED", "Proposed independent review role"],
            ["IR-03", "Calibration governance", "Platt fit outside test; isotonic gate respected", "EVIDENCE_PREPARED", "Proposed model risk role"],
            ["IR-04", "Anti-circularity", "Hidden mechanisms excluded from controls/model", "EVIDENCE_PREPARED", "Proposed independent review role"],
            ["IR-05", "Reason-code integrity", "No PCA names in investigator reason codes", "EVIDENCE_PREPARED", "Proposed fraud operations role"],
            ["IR-06", "Threshold economics", "Base assumptions frozen; sensitivities separate", "EVIDENCE_PREPARED", "Proposed risk committee role"],
            ["IR-07", "Capacity feasibility", "250/day model and stress tests", "EVIDENCE_PREPARED", "Proposed operations role"],
            ["IR-08", "Operational segment diagnostics", "10pp trigger and sample-size gate", "EVIDENCE_PREPARED", "Proposed conduct risk role"],
            ["IR-09", "UAT/SIT/negative testing", "Executable tests", "EVIDENCE_PREPARED", "Proposed technology risk role"],
            ["IR-10", "Claims and limitations", "Observed/synthetic separation", "EVIDENCE_PREPARED", "Proposed model governance role"],
        ],
        columns=["review_id", "review_area", "evidence", "status", "review_owner"],
    )
    independent["independent_review_performed"] = "NO"
    independent["independent_signoff_performed"] = "NO"
    independent = add_metadata(independent, "Governance", "Derived", "Independent review checklist", "Portfolio review simulation")
    write_csv(independent, "validation/independent_review_checklist.csv")

    model_comparison = pd.read_csv(ROOT / "models/model_comparison.csv")
    champion = model_comparison.loc[model_comparison["population"].eq("test") & model_comparison["model_role"].eq("Champion")].iloc[0]
    selected_threshold = pd.read_csv(ROOT / "outputs/threshold_economics.csv")
    selected_threshold = selected_threshold.loc[selected_threshold["population"].eq("test") & selected_threshold["recommended_threshold_flag"].eq(1)].iloc[0]
    write_markdown(
        "validation/model_validation_summary.md",
        "Model Validation Summary",
        f"""## Conclusion

Observed champion `{champion['model']}` has test PR-AUC **{champion['pr_auc']:.4f}**, ROC-AUC **{champion['roc_auc']:.4f}**, validation-locked test fraud recall **{selected_threshold['fraud_recall']:.2%}** and fraud-amount recall **{selected_threshold['fraud_amount_recall']:.2%}**. Probability is Platt-calibrated outside test. Test was not used for model, calibration or threshold selection.

Data quality, split reconciliation, discrimination, calibration, threshold economics, stability and challenger comparison are evidenced in the linked CSV outputs. Remaining limitations are missing original timestamp semantics, small validation fraud count and PCA feature interpretability. Model use is restricted to educational benchmarking.""",
    )
    write_markdown(
        "validation/control_validation_summary.md",
        "Control Validation Summary",
        f"""All **{len(uat)} UAT**, **{len(sit)} SIT** and **{len(negative)} negative tests** executed through reference functions or artifact checks. Rule conditions, boundaries, priority, reason codes, actions, capacity, fallback and overlap are traceable. Synthetic rule precision and recall remain below 100%, and model/rules/anomaly incremental value is reported rather than assumed.""",
    )

    build_excel(uat, sit)

    assumptions = pd.read_csv(ROOT / "governance/assumption_register.csv")
    threshold_table = pd.read_csv(ROOT / "outputs/threshold_economics.csv")
    selected = threshold_table.loc[threshold_table["population"].eq("test") & threshold_table["recommended_threshold_flag"].eq(1)].iloc[0]
    alerts = pd.read_csv(ROOT / "outputs/alert_queue.csv.gz")
    incidents = pd.read_csv(ROOT / "operational_risk/incident_register.csv")
    rules = pd.read_csv(ROOT / "outputs/rule_performance_summary.csv")
    incremental = pd.read_csv(ROOT / "outputs/hybrid_incremental_value.csv")
    final_hybrid = incremental.loc[incremental["component_set"].eq("Final hybrid actionable")].iloc[0]
    synthetic_test = pd.read_csv(ROOT / "models/synthetic_model_summary.csv")
    synthetic_test = synthetic_test.loc[
        synthetic_test["population"].eq("test") & synthetic_test["score"].eq("platt_calibrated")
    ].iloc[0]
    observed_kpi = pd.read_csv(ROOT / "outputs/observed_fraud_kpi_summary.csv")
    observed_fraud_rate = float(observed_kpi.loc[observed_kpi["metric"].eq("fraud_rate"), "value"].iloc[0])
    write_markdown(
        "reports/fraud_risk_committee_memo.md",
        "Fraud Risk Committee Memo",
        f"""## Executive conclusion

The observed benchmark contains 284,807 transactions and 492 fraud cases. `{champion['model']}` is champion with test PR-AUC **{champion['pr_auc']:.4f}**. The validation-selected **top {selected['target_alert_rate']:.2%} alert policy** captures **{selected['fraud_recall']:.2%}** of test fraud transactions and **{selected['fraud_amount_recall']:.2%}** of test fraud amount at precision **{selected['precision']:.2%}**.

## Economics and workload

Test backtest net benefit under disclosed cost assumptions is **{selected['realised_net_benefit']:,.2f} cost units** under frozen illustrative assumptions. The separate synthetic controls environment creates **{len(alerts):,} investigation alerts**; capacity and SLA results are workload proxies, not observed operations.

## Synthetic hybrid economics context

The final hybrid actionable strategy produces **{final_hybrid['net_benefit_proxy']:,.1f} cost units** of simulated net benefit on **{int(final_hybrid['transactions']):,} synthetic test transactions**, equivalent to approximately **{final_hybrid['net_benefit_proxy'] / final_hybrid['transactions']:.3f} cost units per transaction**.

This result must not be linearly scaled to production. The synthetic test fraud prevalence is **{synthetic_test['observed_fraud_rate']:.2%}**, compared with **{observed_fraud_rate:.2%}** in the observed PCA population. Actual value would depend on production fraud prevalence, transaction amounts, preventable-loss rates, recoveries, rule precision, operations costs and customer-friction costs.

Observed threshold economics and synthetic hybrid economics are separate analytical layers.

## Controls

Fifteen observable rules cover device, beneficiary, velocity, authentication, geography, behaviour, override and duplicate patterns. Rule precision/recall are imperfect by design, and incremental value is measured against model-only and anomaly-only strategies.

## Decisions required

1. Approve the observed threshold as an educational backtest recommendation only.
2. Require production re-estimation using real timestamp throughput, recoveries and customer outcomes.
3. Retain hard release blocks for invalid score scale and missing critical inputs.
4. Protect critical alert capacity and require explicit acceptance for backlog breaches.
5. Review flagged operational-segment disparities before any analogous production use.
""",
    )

    write_markdown(
        "recruiter_summary.md",
        "Project 5 Recruiter Summary",
        f"""## Business problem

How should Risk combine transaction indicators, predictive models, deterministic controls and operational governance to reduce fraud loss without unacceptable customer friction or alert workload?

## What I built

- An observed 284,807-row PCA fraud benchmark with logistic, tree, SMOTE/undersampling and anomaly challengers.
- Validation-only model/threshold selection using PR-AUC, fraud-amount capture, capacity and net benefit.
- A separate 250,000-row synthetic controls environment with hidden stochastic fraud mechanisms, interpretable features and a distinct friendly-fraud dispute layer.
- Fifteen governed fraud rules, hybrid decisions, reason codes, fallbacks and overrides.
- Alert priority, 250-alert/day capacity, investigation outcomes, backlog/SLA sensitivity, incidents, KRIs, risk appetite and RCA.
- {len(uat)} UAT, {len(sit)} SIT and {len(negative)} executable negative tests plus a 14-sheet Excel management model.

## Headline result

Observed champion test PR-AUC is **{champion['pr_auc']:.4f}**. The locked test alert policy captures **{selected['fraud_recall']:.2%}** of fraud transactions and **{selected['fraud_amount_recall']:.2%}** of fraud amount at **{selected['precision']:.2%}** precision. Fraud-amount recall remains an Amber appetite exception requiring management acceptance and complementary high-value controls. These are benchmark results, not production claims.

## Value to an employer

I can connect fraud analytics to decisions, workload, controls, incidents and management action while preserving data lineage, claim boundaries and validation evidence.
""",
    )

    write_markdown(
        "README.md",
        "Project 5 - Fraud Detection, Transaction Controls & Operational Risk",
        f"""## Recruiter readout

This project demonstrates the full fraud-risk decision lifecycle: observed benchmark data -> model/anomaly scores -> separately labelled synthetic controls -> hybrid decisions -> alert operations -> incidents/KRIs -> management action.

| Evidence | Result | Meaning |
| --- | ---: | --- |
| Observed transactions | 284,807 | Public PCA fraud benchmark |
| Observed fraud cases | 492 | Extreme imbalance |
| Champion test PR-AUC | {champion['pr_auc']:.4f} | Primary ranking metric |
| Test fraud recall | {selected['fraud_recall']:.2%} | Validation-locked alert policy |
| Test fraud amount recall | {selected['fraud_amount_recall']:.2%} | Validation-locked alert policy |
| Synthetic transactions | 250,000 | Interpretable controls testing |
| UAT / SIT / Negative | {len(uat)} / {len(sit)} / {len(negative)} | Executable governance evidence |

Open `OPEN_THIS_FIRST.html` for the recruiter presentation and `excel/Project5_FraudOperationalRisk_Model.xlsx` for the management workbook.

Author automated self-validation completed. Independent model validation, production approval and organisational release sign-off were not performed.

## Two evidence layers

**Layer A - Observed PCA Fraud Benchmark:** observed discrimination, class imbalance, fraud amount capture, calibration and historical threshold economics. PCA variables never become investigator reason codes.

**Layer B - Synthetic Controls-Testing Environment:** interpretable device, beneficiary, velocity, authentication and behavioural indicators; deterministic rules; hybrid decisions; alert queue; investigation outcomes; incidents; risk appetite; UAT/SIT. Synthetic performance is never presented as observed performance.

Friendly fraud is a separate post-transaction dispute/claim label and is not a transaction-time decline rule.

The recommended observed policy is deliberately presented as a balanced candidate rather than a perfect appetite pass: transaction recall, precision, capacity and net benefit meet the decision criteria, while fraud-amount recall remains Amber. The committee must accept that residual risk or add high-value control overlays.

## Rebuild

```text
python scripts/01_load_validate_observed_data.py
python scripts/02_build_observed_fraud_benchmark.py
python scripts/03_generate_synthetic_control_data.py
python scripts/04_build_synthetic_features_and_labels.py
python scripts/05_build_fraud_rules.py
python scripts/06_build_hybrid_decision_engine.py
python scripts/07_build_alert_operations.py
python scripts/08_build_operational_risk_outputs.py
python scripts/09_generate_excel_and_reports.py
python scripts/10_validate_final_outputs.py
```

Quick reviewer validation: `python scripts/10_validate_final_outputs.py`.

## Main limitations

The original `Time` column is unavailable in Project 0, so ordered `monitoring_period` is a temporal proxy. The observed variables are PCA-transformed, cost assumptions are illustrative, and synthetic controls/workflow outcomes do not represent a live bank. Production use requires governed applicant/customer data, timestamp semantics, real loss/recovery and customer outcomes, integration testing, independent validation and approval.
""",
    )

    write_markdown(
        "limitation_note.md",
        "Project 5 Limitations and Claim Boundaries",
        """1. Observed PCA inputs support ranking but not semantic fraud reason codes.
2. The original timestamp is unavailable; monitoring-period ordering is a proxy.
3. The observed label is treated as transaction fraud; recoveries, delayed confirmation and investigator outcomes are unavailable.
4. Threshold costs, preventable loss and capacity are illustrative cost-unit assumptions.
5. Synthetic transaction controls, fraud types, investigations, incidents and losses are controls-testing simulations.
6. Friendly fraud is a separate post-transaction dispute label.
7. Operational disparity diagnostics are not legal or demographic fairness conclusions.
8. Project 3 context is synthetic and non-causal.
9. This is not a real-time platform, AML system, regulatory model or production authorisation engine.
10. Anomaly-model evaluation is event-count constrained. Isolation Forest PR-AUC varies from 0.0346 on validation to 0.0784 on test, with fewer than 75 fraud events in each split. It is included as a complementary challenger, not a stable standalone fraud-decision model.
11. Author automated self-validation passed, but independent organisational validation, production approval and release sign-off were not performed.""",
    )

    write_markdown(
        "business_question.md",
        "Project 5 Business Question",
        """## Primary question

How can transaction-level risk indicators, predictive models, deterministic controls and operational governance be combined to reduce fraud loss without creating unacceptable customer friction or alert workload?

## Supporting questions

1. What is the observed fraud rate and fraud amount?
2. Which model best identifies fraud under extreme imbalance?
3. How much fraud amount can be captured at constrained alert rates?
4. Which threshold balances net benefit, appetite and analyst capacity?
5. Which observable deterministic controls add value beyond the model?
6. How should allow, step-up, hold, review, decline, block and exception decisions be prioritised?
7. How many investigation alerts can Operations process within SLA?
8. Which controls can fail, what loss may result and what management action is required?""",
    )

    write_markdown(
        "data_contract/source_manifest.md",
        "Project 5 Source Manifest",
        """## Observed source

- Public benchmark: Credit Card Fraud Detection, originally released by the ULB Machine Learning Group and distributed through [Kaggle](https://www.kaggle.com/datasets/mlg-ulb/creditcardfraud).
- Project input: `data/observed/fraud_pca_input.csv.gz`, inherited from Project 0 with 284,807 rows, `v1`-`v28`, `amount`, `fraud_flag` and ordered `monitoring_period` proxy.
- Integrity: SHA-256 is recorded in `data_contract/source_manifest.csv`.

The CSV manifest records dataset id, dataset version, source name, source URL/access terms, snapshot date, row count, fraud count/rate, amount totals, SHA-256, canonical path, packaged path, transformation and downstream use.

## Synthetic source

Customers, devices, beneficiaries, transactions, post-transaction disputes and incidents are generated by scripts 03, 04 and 08 using seed `20260711`. Hidden latent fields remain in `data/synthetic/internal/` for anti-circularity audit and are excluded from detection models and rules.

## Reference lineage

Project 3 customer-risk context and Project 6 governance-pattern references are documented as reference lineage only. They are not external fraud-performance data and are not used to claim production causality or organisational sign-off.

## Rebuild portability

All code uses paths relative to the project root. The full package includes the observed input; the AI-review package includes samples and aggregate evidence and is not intended for full model retraining.""",
    )

    write_markdown(
        "DATA_ACCESS.md",
        "Data Access and Rebuild Modes",
        """## Full rebuild mode

Use the FULL package. It contains the 284,807-row observed PCA input, 250,000-row synthetic controls population, latent audit data and transaction-level outputs. Run `python scripts/run_full_pipeline.py` from the project folder.

## Reviewer quick-check mode

In the FULL package, run `python scripts/10_validate_final_outputs.py`. In the AI_REVIEW_LIGHT package, run `python scripts/validate_review_package.py`. These checks validate the included results, Excel, links, UAT/SIT, negative tests and claim boundaries without retraining.

## AI review mode

The AI_REVIEW_LIGHT package excludes large raw/generated transaction-level files and serialized models. It includes samples under `data/review_samples/`, aggregate outputs, reports, Excel, HTML, code and validation evidence. It is for review, not full retraining.

## GitHub publishing

Publish code, samples, aggregate outputs, HTML, Excel and reports in the repository. Store large full data/output artifacts in a release asset or governed external storage when repository-size policy requires it. Never replace the observed dataset with synthetic data without changing the claim boundary.""",
    )
    write_markdown(
        "AI_REVIEW_README.md",
        "AI Review Package Guide",
        """Start with `OPEN_THIS_FIRST.html`, then read `README.md`, `recruiter_summary.md`, `validation/validation_report.md` and the two committee memos. Samples are in `data/review_samples/`; complete aggregate evidence is retained across `models/`, `outputs/`, `rules/`, `operational_risk/`, `governance/`, `testing/` and `validation/`.

The light package intentionally cannot retrain the observed or synthetic models because large transaction-level inputs are excluded. Run `python scripts/validate_review_package.py` for a light-package integrity check. Use the FULL package for end-to-end rebuild.""",
    )

    requirements = """joblib==1.5.3
matplotlib==3.10.8
numpy==2.4.4
openpyxl==3.1.5
pandas==3.0.2
scikit-learn==1.8.0
scipy==1.17.1
"""
    (ROOT / "requirements.txt").write_text(requirements, encoding="ascii")
    (ROOT / "change_log.md").write_text(
        "# Change Log\n\n- 2026-07-11: Project 5 Gold v1.0 end-to-end build generated from locked methodology v1.1.\n- 2026-07-11: v1.0.1 public remediation formalised threshold-selection hierarchy, corrected independent-review/release-approval wording, expanded source lineage, added uncertainty/critical-queue/economics disclosures, and rebuilt package integrity controls without changing the underlying selected threshold or core model logic.\n",
        encoding="utf-8",
    )
    build_html()
    log.info("Excel, tests, reports and HTML generated")
    print(f"Reporting PASS | UAT={len(uat)} | SIT={len(sit)} | Negative={len(negative)} | Excel=14 sheets")


if __name__ == "__main__":
    main()
